import pandas as pd
import openpyxl
import os
import sys
from datetime import datetime

# ─────────────────────────────────────────────────────────
# SOLUTION: Find users in activity log but
#           missing from active employee list
#
# Input: ONE Excel file with TWO sheets:
#   Sheet 1 — Activity Log     (user_id, login_time, action, department...)
#   Sheet 2 — Active Employees (emp_id, name, department, status...)
#
# Memory safe: openpyxl read_only streaming for any file size
# ─────────────────────────────────────────────────────────

EXCEL_FILE  = "data/employee_activity.xlsx"
OUTPUT_FILE = "missing_users_report.xlsx"
CHUNK_SIZE  = 50000


def fix_value(val):
    """Convert Excel serial date numbers or datetime objects to clean strings."""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, float) and 40000 < val < 60000:
        return (pd.Timestamp('1899-12-30') +
                pd.Timedelta(days=val)).strftime('%Y-%m-%d %H:%M:%S')
    return val


def get_sheet_names(filepath):
    """Get all sheet names from Excel file."""
    wb     = openpyxl.load_workbook(filepath, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def detect_id_column(columns, file_label):
    """Auto-detect the ID column from common names."""
    common = ['user_id', 'emp_id', 'empid', 'employee_id',
              'userid', 'EmpID', 'id', 'staff_id']
    col = next((c for c in columns if c.strip().lower()
                in [x.lower() for x in common]), None)
    if not col:
        print(f"  Columns in {file_label}: {list(columns)}")
        col = input(f"  Enter ID column name for {file_label}: ").strip()
    print(f"  ID column in {file_label}: '{col}'")
    return col


def load_active_employees(filepath, sheet_name):
    """Load active employee IDs from the employee sheet."""
    print(f"Loading active employees from sheet: '{sheet_name}'...")
    df         = pd.read_excel(filepath, sheet_name=sheet_name)
    id_col     = detect_id_column(df.columns, sheet_name)
    active_ids = set(df[id_col].astype(str).str.strip())
    print(f"  Total active employees: {len(active_ids):,}")
    return active_ids


def stream_sheet_chunks(filepath, sheet_name, chunksize):
    """
    TRUE streaming of a specific sheet using openpyxl read_only.
    Reads row by row — never loads full sheet into memory.
    Safe for any file size.
    """
    wb  = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws  = wb[sheet_name]
    rows    = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]

    batch = []
    for row in rows:
        fixed_row = tuple(fix_value(v) for v in row)
        batch.append(fixed_row)
        if len(batch) == chunksize:
            yield pd.DataFrame(batch, columns=headers)
            batch = []
    if batch:
        yield pd.DataFrame(batch, columns=headers)

    wb.close()


def find_missing_users(filepath, activity_sheet, active_ids):
    """Find all user IDs in activity log not in employee list."""
    print(f"\nProcessing '{activity_sheet}' in chunks of {CHUNK_SIZE:,} rows...")

    common        = ['user_id', 'emp_id', 'empid', 'employee_id', 'userid', 'EmpID']
    user_id_col   = None
    missing_users = set()
    total_rows    = 0
    chunk_num     = 0

    for chunk in stream_sheet_chunks(filepath, activity_sheet, CHUNK_SIZE):
        chunk_num += 1

        if user_id_col is None:
            user_id_col = next((c for c in chunk.columns if c.strip().lower()
                                in [x.lower() for x in common]), None)
            if not user_id_col:
                print(f"  Columns: {list(chunk.columns)}")
                user_id_col = input("  Enter user ID column name: ").strip()
            print(f"  Auto-detected ID column: '{user_id_col}'")

        total_rows += len(chunk)
        chunk_ids   = set(chunk[user_id_col].astype(str).str.strip())
        missing_users.update(chunk_ids - active_ids)

        print(f"  Chunk {chunk_num:>3}: {total_rows:>9,} rows processed | "
              f"Ghost users found: {len(missing_users)}")

    return missing_users, total_rows, user_id_col


def generate_report(filepath, activity_sheet, missing_users, user_id_col, output_file):
    """Build clean Excel report of ghost user activity."""
    print("\nGenerating final report...")

    chunks = []
    for chunk in stream_sheet_chunks(filepath, activity_sheet, CHUNK_SIZE):
        chunk[user_id_col] = chunk[user_id_col].astype(str).str.strip()
        filtered = chunk[chunk[user_id_col].isin(missing_users)]
        if not filtered.empty:
            chunks.append(filtered)

    missing_df = pd.concat(chunks, ignore_index=True)

    date_col = next((c for c in missing_df.columns if any(
                     k in c.lower() for k in ['time', 'date'])), None)
    dept_col = next((c for c in missing_df.columns if any(
                     k in c.lower() for k in ['dept', 'department'])), None)

    rows = []
    for uid, group in missing_df.groupby(user_id_col):
        row = {'User ID': uid, 'Total Actions': len(group)}
        if dept_col:
            row['Departments Accessed'] = ', '.join(sorted(group[dept_col].unique()))
        if date_col:
            row['First Activity'] = str(group[date_col].min())
            row['Last Activity']  = str(group[date_col].max())
        rows.append(row)

    summary = pd.DataFrame(rows).sort_values('Total Actions', ascending=False)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name='Missing Users Summary', index=False)
        missing_df.to_excel(writer, sheet_name='Full Activity Detail', index=False)

    print(f"  Report saved: {output_file}")
    return summary


def main():
    print("=" * 60)
    print("   EMPLOYEE ACTIVITY ANALYSIS — MISSING USER FINDER")
    print("=" * 60)

    if not os.path.exists(EXCEL_FILE):
        print(f"\nERROR: File not found → {EXCEL_FILE}")
        sys.exit(1)

    file_mb = os.path.getsize(EXCEL_FILE) / (1024 * 1024)
    print(f"\nFile: {EXCEL_FILE} ({file_mb:.1f} MB)")

    # Show available sheets
    sheets = get_sheet_names(EXCEL_FILE)
    print(f"Sheets found: {sheets}")

    # Assume Sheet 1 = activity, Sheet 2 = employees
    activity_sheet = sheets[0]
    employee_sheet = sheets[1]
    print(f"  Activity sheet  : '{activity_sheet}'")
    print(f"  Employee sheet  : '{employee_sheet}'")

    active_ids                         = load_active_employees(EXCEL_FILE, employee_sheet)
    missing_users, total_rows, uid_col = find_missing_users(EXCEL_FILE, activity_sheet, active_ids)
    summary                            = generate_report(EXCEL_FILE, activity_sheet,
                                                         missing_users, uid_col, OUTPUT_FILE)

    print("\n" + "=" * 60)
    print(f"  TOTAL ROWS PROCESSED  : {total_rows:,}")
    print(f"  ACTIVE EMPLOYEES      : {len(active_ids):,}")
    print(f"  MISSING / GHOST USERS : {len(missing_users)}")
    print("=" * 60)
    print("\nMissing Users Summary:")
    print(summary.to_string(index=False))
    print(f"\nFull report saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()